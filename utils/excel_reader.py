from openpyxl import load_workbook
from datetime import datetime

# Utility function to read login credentials from Excel sheet
def read_login_data(file_path):
    wb = load_workbook(file_path)
    sheet = wb.active

    data = []

    for row in range(2, sheet.max_row + 1): # Iterate through all test data rows
        username = sheet.cell(row=row, column=5).value  # Column E
        password = sheet.cell(row=row, column=6).value  # Column F

        if username and password: # Store username and password for execution
            data.append((row, username, password))

    return data, wb, sheet


def write_result(wb, sheet, row, result):
# Write execution result back to Excel
    now = datetime.now()

    sheet.cell(row=row, column=3).value = now.date()  # Date column
    sheet.cell(row=row, column=7).value = result     # Result column

    wb.save("test_data/Task_15 (1).xlsx")